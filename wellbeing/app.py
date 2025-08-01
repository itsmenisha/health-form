# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook

import os


app = Flask(__name__)

# === EXCEL FILES PATH ===
DATA_PATH = 'data'
os.makedirs(DATA_PATH, exist_ok=True)

files = {
    'who5': os.path.join(DATA_PATH, 'who5.xlsx'),
    'gad7': os.path.join(DATA_PATH, 'gad7.xlsx'),
    'phq9': os.path.join(DATA_PATH, 'phq9.xlsx'),
}

# === CREATE IF NOT EXIST ===
for key, path in files.items():
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Address', 'Age', 'Date',
                  'Total Score', 'Percent Score'])
        wb.save(path)

# === ROUTES ===


@app.route('/')
def frontpage():
    return render_template('frontpage.html')


@app.route('/navbar')
def navbar():
    return render_template('navbar.html')


@app.route('/support')
def support():
    return render_template('supportsystem.html')


@app.route('/about')
def about():
    return render_template('about.html')


@app.route('/services')
def services():
    return render_template('services.html')


@app.route('/faq')
def faq():
    return render_template('faq.html')


@app.route('/screen_info')
def screen():
    return render_template('screening.html')


@app.route('/who5')
def who5():
    return render_template('forms/who5.html')


@app.route('/gad7')
def gad7():
    return render_template('forms/gad7.html')


@app.route('/phq9')
def phq9():
    return render_template('forms/phq9.html')


@app.route('/introwho')
def introwho():
    return render_template('forms/whodetails.html')


@app.route('/introgad')
def introgad():
    return render_template('forms/gaddetails.html')


@app.route('/introphq')
def introphq():
    return render_template('forms/phqdetails.html')


@app.route('/article1')
def article1():
    return render_template('article1.html')


@app.route('/article2')
def article2():
    return render_template('article2.html')


@app.route('/article3')
def article3():
    return render_template('article3.html')


@app.route('/photo/<path:filename>')
def photo(filename):
    return send_from_directory('photo', filename)
# === HANDLE FORM SUBMISSIONS ===


@app.route('/submit/<form_type>', methods=['POST'])
def submit(form_type):
    if form_type not in files:
        return jsonify({"message": "Unknown form type"}), 400

    data = request.get_json()
    name = data.get('name')
    address = data.get('address')
    age = data.get('age')
    date = data.get('date')
    total_Score = data.get('totalScore')
    percent_score = data.get('percentScore')

    try:
        wb = load_workbook(files[form_type])
        ws = wb.active
        ws.append([name, address, age, date, total_Score, percent_score])
        wb.save(files[form_type])
        return jsonify({"message": "✅ We will calculate the data."})
    except Exception as e:
        return jsonify({"message": f"❌ Failed to save: {e}"}), 500


if __name__ == '__main__':
    app.run(debug=True)
